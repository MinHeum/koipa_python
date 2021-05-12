import openpyxl as xl

exf =  xl.load_workbook('c:\\dd\\itx.xlsx')
aws = exf.active
tot = 0
for i in aws.rows:
    index = i[0].row
    name = i[0].value
    salary = i[1].value
    tot += salary

last_ros = len(tuple(aws.rows))+1
aws.cell(row = last_ros, column = 1).value = "평균"
aws.cell(row = last_ros, column = 2).value = tot / len(tuple(aws.rows))


exf.save('c:\\dd\\itxout.xlsx')
exf.close()
